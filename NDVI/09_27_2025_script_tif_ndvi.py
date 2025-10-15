#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Sun Sep 28 15:22:11 2025

@author: melis
"""

import xarray as xr
import rasterio
import rasterio.mask
import geopandas as gpd
import numpy as np
from rasterio.transform import from_origin
from pathlib import Path
import pandas as pd
import json

z = 2022

def extract_and_filter_ndvi(nc_path, output_path, boundary_path, var_name="NDVI", 
                            min_ndvi=-1.0, max_ndvi=1.0):
    """
    Extract NDVI from .nc file, crop to California boundary, filter invalid values, and save as GeoTIFF.
    """
    ds = xr.open_dataset(nc_path)
    if var_name not in ds:
        raise ValueError(f"{var_name} not found in {nc_path}. Available: {list(ds.keys())}")
    
    ndvi = ds[var_name].squeeze()
    data = ndvi.values.astype("float32")

    # --- Handle latitude/longitude naming ---
    if "lat" in ds:
        lat = ds["lat"].values
    elif "latitude" in ds:
        lat = ds["latitude"].values
    else:
        raise KeyError("No latitude variable found in dataset")

    if "lon" in ds:
        lon = ds["lon"].values
    elif "longitude" in ds:
        lon = ds["longitude"].values
    else:
        raise KeyError("No longitude variable found in dataset")

    # Build transform
    transform = from_origin(lon.min(), lat.max(), np.abs(lon[1]-lon[0]), np.abs(lat[1]-lat[0]))

    # Save raw NDVI temporarily to apply exact mask
    tmp_meta = {
        "driver": "GTiff",
        "height": data.shape[0],
        "width": data.shape[1],
        "count": 1,
        "dtype": "float32",
        "crs": "EPSG:4326",   # VIIRS NDVI is in lat/lon WGS84
        "transform": transform,
        "nodata": np.nan
    }

    tmp_tif = str(output_path).replace(".tif", "_tmp.tif")
    with rasterio.open(tmp_tif, "w", **tmp_meta) as dst:
        dst.write(data, 1)

    # --- Apply California shapefile mask ---
    boundary = gpd.read_file(boundary_path).to_crs("EPSG:4326")
    shapes = [json.loads(boundary.to_json())["features"][0]["geometry"]]
    
    
    
    with rasterio.open(tmp_tif) as src:
        masked_data, masked_transform = rasterio.mask.mask(src, shapes, crop=True, filled=True, nodata=np.nan)
        band = masked_data[0]  # extract the single band, shape = (H, W)

        masked_meta = src.meta.copy()
        masked_meta.update({
        "count": 1,
        "height": band.shape[0],
        "width": band.shape[1],
        "transform": masked_transform,
        "nodata": np.nan,
        "dtype": "float32"
        })
        # masked_data shape = (1, height, width)
    

    # --- Stats before filtering ---
    total_pixels = band.size
    nan_pixels = np.sum(np.isnan(band))
    valid_pixels = total_pixels - nan_pixels
    print(f"\n=== {nc_path.name} (Cropped to CA) ===")
    print(f"Total pixels: {total_pixels:,}")
    print(f"NaN pixels: {nan_pixels:,} ({nan_pixels/total_pixels*100:.2f}%)")
    print(f"Non-NaN pixels: {valid_pixels:,}")
    if valid_pixels > 0:
        print(f"NDVI range: {np.nanmin(band):.4f} to {np.nanmax(band):.4f}")
        print(f"Mean NDVI: {np.nanmean(band):.4f}")

    # --- Filtering: outside (-1, 1) → NaN ---
    invalid_range = (band <= min_ndvi) | (band >= max_ndvi)
    range_removed = np.sum(invalid_range & ~np.isnan(band))
    filtered = np.where(invalid_range, np.nan, band)

    # --- Stats after filtering ---
    final_valid = np.sum(~np.isnan(filtered))
    print("\n--- Filtering Results ---")
    print(f"Pixels removed (outside range ({min_ndvi}, {max_ndvi})): {range_removed:,}")
    print(f"Final valid pixels: {final_valid:,} ({final_valid/total_pixels*100:.2f}%)")
    if final_valid > 0:
        print(f"Filtered NDVI range: {np.nanmin(filtered):.4f} to {np.nanmax(filtered):.4f}")
        print(f"Filtered mean NDVI: {np.nanmean(filtered):.4f}")

    # Save cropped + filtered GeoTIFF
    with rasterio.open(output_path, "w", **masked_meta) as dst:
        dst.write(filtered.astype("float32"), 1)

    print(f"✅ Saved CA-cropped NDVI to {output_path}")
    Path(tmp_tif).unlink(missing_ok=True)  # remove temp file

    return output_path


def process_ndvi_folder(input_dir, output_dir, boundary_path, var_name="NDVI"):
    """
    Process all NDVI .nc files in a folder:
    - Extract NDVI
    - Crop to California boundary
    - Filter invalid values
    - Save single-band GeoTIFF with renamed file
    - Detect missing days (per year)
    """
    input_dir = Path(input_dir)
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    nc_files = sorted(input_dir.glob("*.nc"))
    if not nc_files:
        print(f"❌ No .nc files found in {input_dir}")
        return

    print(f"🔍 Found {len(nc_files)} NDVI files to process...\n")

    dates_found = []
    for nc_file in nc_files:
        # Extract YYYYMMDD from filename
        parts = nc_file.stem.split("_")
        date_str = [p for p in parts if p.isdigit() and len(p) == 8][0]  # "20240523"
        date = pd.to_datetime(date_str, format="%Y%m%d")
        dates_found.append(date)

        # New filename format
        out_name = f"NDVI_{date.strftime('%m_%d_%Y')}.tif"
        out_path = output_dir / out_name

        # Process file
        extract_and_filter_ndvi(nc_file, out_path, boundary_path, var_name=var_name)

    # --- Missing days check per year ---
    all_dates = pd.Series(sorted(dates_found))
    if not all_dates.empty:
        print("\n=== Missing Days Report (per year) ===")
        for year, group in all_dates.groupby(all_dates.dt.year):
            year_dates = sorted(group)
            start = pd.Timestamp(f"{year}-01-01")
            end = pd.Timestamp(f"{year}-12-31")
            full_range = pd.date_range(start, end, freq="D")
            missing = full_range.difference(group)

            print(f"\nYear {year}:")
            print(f"  Total days in year: {len(full_range)}")
            print(f"  Available NDVI days: {len(group)}")
            print(f"  Missing days: {len(missing)}")
            if len(missing) > 0:
                print("  Missing dates:")
                for d in missing:
                    print("   -", d.strftime("%Y-%m-%d"))
    else:
        print("⚠️ No valid NDVI dates parsed.")
    return all_dates


# === MAIN EXECUTION ===
if __name__ == "__main__":
    input_dir = fr"/Users/melis/Desktop/NDVI_Download_09_22_2025/ndvi_data_{z}"
    output_dir = fr"/Users/melis/Desktop/NDVI_Download_09_22_2025/ndvi_data_{z}_cropped_single_layer"
    boundary_path = "/Users/melis/Downlaods/hilsside_only _figure/only_hillside_landslide/Fixed_geometries/CA_Boundary/California_Boundary.shp"
    
    
    all_dates = process_ndvi_folder(input_dir, output_dir, boundary_path)
    # Save missing-days report to CSV
    report_path ="missing_days_report.csv"
    rows = []
    for year, group in all_dates.groupby(all_dates.dt.year):
        start = pd.Timestamp(f"{year}-01-01")
        end = pd.Timestamp(f"{year}-12-31")
        full_range = pd.date_range(start, end, freq="D")
        missing = full_range.difference(group)
    
        for d in missing:
            rows.append({"Year": year, "MissingDate": d.strftime("%Y-%m-%d")})
    
    if rows:
        pd.DataFrame(rows).to_csv(report_path, index=False)
        print(f"\n📂 Missing days report saved to {report_path}")
    else:
        print("\n✅ No missing days detected, nothing written to CSV")


#%% Convert to full year excel

import rasterio
import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime

# === Input directory with NDVI GeoTIFFs ===
input_dir = Path(fr"/Users/melis/Desktop/NDVI_Download_09_22_2025/ndvi_data_{z}_cropped_single_layer")
excel_path = input_dir / fr"NDVI_{z}_AllDays.xlsx"

# === Collect and sort .tiff files ===
tif_files = sorted(
    input_dir.glob("*.tif"),
    key=lambda p: datetime.strptime("_".join(p.stem.split("_")[1:]), "%m_%d_%Y")
)

print(f"Found {len(tif_files)} NDVI files")

dfs = []
day_labels = []

for day_idx, tif_path in enumerate(tif_files):
    parts = tif_path.stem.split("_")
    date_str = "_".join(parts[1:])  # e.g. "03_28_2018"
    dt = datetime.strptime(date_str, "%m_%d_%Y")

    col_label = str(day_idx)  # 0, 1, 2, ...

    with rasterio.open(tif_path) as src:
        ndvi = src.read(1)
        transform = src.transform
        nodata = src.nodata

        nrows, ncols = ndvi.shape
        rows, cols = np.meshgrid(np.arange(nrows), np.arange(ncols), indexing="ij")

        ndvi_flat = ndvi.flatten()
        rows_flat = rows.flatten()
        cols_flat = cols.flatten()

        # Compute bounding box (no rounding)
        x_min, y_min = transform * (cols_flat, rows_flat)
        x_max, y_max = transform * (cols_flat + 1, rows_flat + 1)

        # Filter valid pixels
        valid_mask = np.isfinite(ndvi_flat)
        if nodata is not None:
            valid_mask &= ndvi_flat != nodata

        df = pd.DataFrame({
            "x_min": x_min[valid_mask],
            "y_min": y_min[valid_mask],
            "x_max": x_max[valid_mask],
            "y_max": y_max[valid_mask],
            col_label: ndvi_flat[valid_mask]
        })

        dfs.append(df)
        day_labels.append(col_label)

# === Merge all days ===
print("Merging all daily NDVI values...")
master_df = dfs[0]
for i in range(1, len(dfs)):
    master_df = pd.merge(
        master_df,
        dfs[i],
        on=["x_min", "y_min", "x_max", "y_max"],
        how="outer"
    )

# Add NDVI_Index
master_df.insert(0, "NDVI_Index", np.arange(1, len(master_df) + 1))

# === Save to Excel ===
master_df.to_excel(excel_path, index=False)
print(f"✅ Excel saved: {excel_path}")

#%% Single day Excel file

import rasterio
import pandas as pd
import numpy as np
from pathlib import Path

# === Input file ===
tif_path = Path("/Users/melis/Desktop/NDVI_Download_09_22_2025/ndvi_data_{z}_cropped_single_layer/NDVI_01_01_{z}.tif")

# === Extract date string from filename ===
# Split by "_" and pick the element that looks like YYYYMMDD
parts = tif_path.stem.split("_")
date_str = "_".join(parts[1:])  # "01_01_2018"

excel_path = tif_path.parent / f"NDVI_{date_str}.xlsx"

# === Open the raster ===
with rasterio.open(tif_path) as src:
    ndvi = src.read(1)  # NDVI values
    transform = src.transform
    nodata = src.nodata

    # Dimensions
    nrows, ncols = ndvi.shape

    # Create row/col index grids
    rows, cols = np.meshgrid(np.arange(nrows), np.arange(ncols), indexing="ij")

    # Flatten arrays
    ndvi_flat = ndvi.flatten()
    rows_flat = rows.flatten()
    cols_flat = cols.flatten()

    # Compute pixel bounds (x_min, y_min, x_max, y_max)
    x_min, y_min = transform * (cols_flat, rows_flat)
    x_max, y_max = transform * (cols_flat + 1, rows_flat + 1)

    # Mask out invalid values (nodata or NaN)
    valid_mask = np.isfinite(ndvi_flat)
    if nodata is not None:
        valid_mask &= ndvi_flat != nodata

    # Build dataframe with only valid pixels
    df = pd.DataFrame({
        "x_min": x_min[valid_mask],
        "y_min": y_min[valid_mask],
        "x_max": x_max[valid_mask],
        "y_max": y_max[valid_mask],
        "NDVI_Value": ndvi_flat[valid_mask]
    })

    # Reset NDVI_Index to 1..N
    df.insert(0, "NDVI_Index", np.arange(1, len(df) + 1))

# === Save to Excel ===
df.to_excel(excel_path, index=False)

print(f"Excel saved: {excel_path}")

#%% Create common excel includes both merged and daily

#%% Missing columns

from datetime import datetime
z = 2022
excel_in =  fr"/Users/melis/Desktop/NDVI_Download_09_22_2025/ndvi_data_{z}_cropped_single_layer/NDVI_{z}_AllDays.xlsx"
excel_out = fr"/Users/melis/Desktop/NDVI_Download_09_22_2025/ndvi_data_{z}_cropped_single_layer/NDVI_{z}_AllDays_filled.xlsx"

def date_to_day_number(date_str, start_date="2022-01-01"):
    """
    Convert a date string (YYYY-MM-DD) into a day number offset from start_date.
    Example: start_date = '2022-01-01' → day 0
    """
    # Parse both as datetime
    d = datetime.strptime(date_str, "%Y-%m-%d")
    s = datetime.strptime(start_date, "%Y-%m-%d")
    
    # Compute offset
    delta = (d - s).days
    return delta

missing_dates = ("2022-07-27", "2022-07-28", "2022-07-29", "2022-07-30", "2022-07-31", "2022-08-01", "2022-08-02", "2022-08-03", "2022-08-04", "2022-08-05", "2022-08-06", "2022-08-07", "2022-08-08", "2022-08-09", "2022-08-10")
missing_day_nums = [date_to_day_number(d) for d in missing_dates]
print("Missing day numbers: ", missing_day_nums)

#==== Load Excel =====
df = pd.read_excel(excel_in)

#ensure first 5 columns

meta_cols = df.iloc[:, :5]
day_columns = df.iloc[:, 5:]

#rebuild a dataframe with 365 columns

new_days = pd.DataFrame(index = day_columns.index)

for day in range(365):
    if day in missing_day_nums:
        # Insert empty column for missing day
        new_days[str(day)] = pd.NA
        
    else:
        #Use existing column(shifted because of missing days)
        new_days[str(day)] = day_columns. iloc[:, 0]
        day_columns = day_columns.iloc[:, 1:] #drop frist column each time
    
#merge back the metadata and +days
df_filled = pd.concat([meta_cols, new_days], axis =1)
#Save new excel
df_filled.to_excel(excel_out, index=False)
print(f"Saved filled Excel with missing days inserted --> {excel_out}")

#%% Compare filled and not filled excels

import pandas as pd
z = 2022

# Paths
filled_path = fr"/Users/melis/Desktop/NDVI_Download_09_22_2025/ndvi_data_{z}_cropped_single_layer/NDVI_{z}_AllDays.xlsx"
not_filled_path = fr"/Users/melis/Desktop/NDVI_Download_09_22_2025/ndvi_data_{z}_cropped_single_layer/NDVI_{z}_AllDays_filled.xlsx"

# Load only first 5 columns (metadata + keys)
usecols = list(range(5))
df_filled = pd.read_excel(filled_path, usecols=usecols)
df_not_filled = pd.read_excel(not_filled_path, usecols=usecols)

# Pick key columns (x_min, y_min, x_max, y_max → columns 1:5)
key_cols = df_filled.columns[1:5]

# Build key tuples
df_filled["key"] = df_filled[key_cols].astype(str).agg("_".join, axis=1)
df_not_filled["key"] = df_not_filled[key_cols].astype(str).agg("_".join, axis=1)

keys_filled = set(df_filled["key"])
keys_not_filled = set(df_not_filled["key"])

# Differences
missing_in_filled = keys_not_filled - keys_filled
missing_in_not_filled = keys_filled - keys_not_filled

print("Unique keys in filled:", len(keys_filled))
print("Unique keys in not filled:", len(keys_not_filled))
print("Missing in filled:", len(missing_in_filled))
print("Missing in not filled:", len(missing_in_not_filled))

# Save mismatched rows for inspection
if missing_in_filled:
    df_not_filled[df_not_filled["key"].isin(missing_in_filled)].to_csv("rows_missing_in_filled.csv", index=False)
if missing_in_not_filled:
    df_filled[df_filled["key"].isin(missing_in_not_filled)].to_csv("rows_missing_in_not_filled.csv", index=False)

print("✅ Saved mismatched rows as CSV")

#Find the duplicates

from openpyxl import load_workbook

filled_path = fr"/Users/melis/Desktop/NDVI_Download_09_22_2025/ndvi_data_2022_cropped_single_layer/NDVI_2022_AllDays_filled.xlsx"
not_filled_path = fr"/Users/melis/Desktop/NDVI_Download_09_22_2025/ndvi_data_2022_cropped_single_layer/NDVI_2022_AllDays.xlsx"


def count_rows(path):
    wb = load_workbook(path, read_only=True)
    ws = wb.active
    nrows = 0
    last_nonempty = 0
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if any(cell is not None and str(cell).strip() != "" for cell in row):
            last_nonempty = i
        nrows += 1
    return nrows, last_nonempty

print("Filled:", count_rows(filled_path))
print("Not filled:", count_rows(not_filled_path))

#%% Sanity check for 2022 filled and non filled

import random
# Paths
filled_path = fr"/Users/melis/Desktop/NDVI_Download_09_22_2025/ndvi_data_{z}_cropped_single_layer/NDVI_{z}_AllDays.xlsx"
not_filled_path = fr"/Users/melis/Desktop/NDVI_Download_09_22_2025/ndvi_data_{z}_cropped_single_layer/NDVI_{z}_AllDays_filled.xlsx"


df_filled = pd.read_excel(filled_path)
df_not_filled = pd.read_excel(not_filled_path)

#Check column names for both

filled_cols = list(df_filled.columns)
not_filled_cols = list(df_not_filled.columns)

if filled_cols == not_filled_cols:
    print("✅ Columns are in the same order and identical.")
else:
    print("⚠️ Columns differ.")

# Check columns that exist in one but not the other
extra_in_filled = set(filled_cols) - set(not_filled_cols)
extra_in_not_filled = set(not_filled_cols) - set(filled_cols)

print("Columns only in filled file:", extra_in_filled)
print("Columns only in not-filled file:", extra_in_not_filled)


#pick random row/col indices

import random

n_checks = 10 
rows = random.sample(range(min(len(df_filled), len(df_not_filled))), n_checks)

n_col_checks = min(5, len(filled_cols))  
cols = random.sample(range(len(filled_cols)), n_col_checks)

for r in rows:
    print(f"\nRow {r}")
    for c in cols:
        colname = filled_cols[c]
        val_filled = df_filled.iloc[r, c]      # ✅ correct way
        val_not_filled = df_not_filled.iloc[r, c]  # ✅ same here
        print(f" {colname}: filled = {val_filled}, not_filled = {val_not_filled}")
        
#%% Check column names in all years
z = {2017,2018,2019,2020,2021,2022,2023,2024}
import pandas as pd

files = {
    2017: r"/Users/melis/Desktop/NDVI_Download_09_22_2025/ndvi_data_2017_cropped_single_layer/NDVI_2017_AllDays.xlsx",
    2018: r"/Users/melis/Desktop/NDVI_Download_09_22_2025/ndvi_data_2018_cropped_single_layer/NDVI_2018_AllDays.xlsx",
    2019: r"/Users/melis/Desktop/NDVI_Download_09_22_2025/ndvi_data_2019_cropped_single_layer/NDVI_2019_AllDays.xlsx",
    2020: r"/Users/melis/Desktop/NDVI_Download_09_22_2025/ndvi_data_2020_cropped_single_layer/NDVI_2020_AllDays.xlsx",
    2021: r"/Users/melis/Desktop/NDVI_Download_09_22_2025/ndvi_data_2021_cropped_single_layer/NDVI_2021_AllDays.xlsx",
    2022: r"/Users/melis/Desktop/NDVI_Download_09_22_2025/ndvi_data_2022_cropped_single_layer/NDVI_2022_AllDays.xlsx",
    2023: r"/Users/melis/Desktop/NDVI_Download_09_22_2025/ndvi_data_2023_cropped_single_layer/NDVI_2023_AllDays.xlsx",
    2024: r"/Users/melis/Desktop/NDVI_Download_09_22_2025/ndvi_data_2024_cropped_single_layer/NDVI_2024_AllDays.xlsx",
}

all_cols = {}
for year, path in files.items():
    df = pd.read_excel(path)
    all_cols[year] = list(df.columns)
    print(f"{year}: {len(all_cols[year])} columns")

# Example: check differences between years
print("\nColumns in 2017 but not in 2018:")
print(set(all_cols[2017]) - set(all_cols[2023]))


#%% Count number of rows in each file

import pandas as pd
from pathlib import Path

# Base folder
base = Path("/Users/melis/Desktop/09_30_ndvi_excels")

# Years to check
years = range(2017, 2025)

row_counts = {}
col_counts = {}

for z in years:
    file = base /  f"NDVI_{z}_AllDays.xlsx"
    
    # Read only row count
    df = pd.read_excel(file, engine="openpyxl")
    
    row_counts[z] = len(df)
    col_counts[z] = len(df.columns)
    
    print(f"Year {z}: rows = {row_counts[z]:,}, columns = {col_counts[z]}")

# Summary
print("\n=== Summary ===")
for z in years:
    print(f"{z}: {row_counts[z]:,} rows, {col_counts[z]} columns")

#%% Find the unique keys in excel 2021

import pandas as pd
from pathlib import Path

#==== Load only the bounding box coordinates

xlsx_path = Path(r"/Users/melis/Desktop/09_30_ndvi_excels/myadjustmenetmanually/NDVI_2021_AllDays.xlsx")
df = pd.read_excel(xlsx_path, usecols ="B:E")

#=====Create Dictionary ======

bbox_dict = {tuple(row): idx+1 for idx, row in df.iterrows()}

print("Total rows: ", len(bbox_dict))
print("First 5 entries: ")

for k,v in list(bbox_dict.items())[:5]:
    print(k, "==>", v)
    
# =====Check for duplicates ======

bbox_tuples = list(df.itertuples(index=False, name = None))

#Count duplicates

total = len(bbox_tuples)
unique = len(set(bbox_tuples))
duplicates = total - unique

import json
from decimal import Decimal, getcontext, ROUND_HALF_UP

# --------------------- CONFIG ---------------------

# How many decimals to keep in keys (string format).
# With your grid (~0.05°), 6 is conservative and robust.
DECIMALS = 10

# Change these paths to yours:
BASE_2021 = Path(r"/Users/melis/Desktop/09_30_ndvi_excels/myadjustmenetmanually/NDVI_2021_AllDays.xlsx")

OTHER_YEARS = {
    # year: path
    2017: Path(r"/Users/melis/Desktop/09_30_ndvi_excels/myadjustmenetmanually/NDVI_2017_AllDays.xlsx"),
    2020: Path(r"/Users/melis/Desktop/09_30_ndvi_excels/myadjustmenetmanually/NDVI_2020_AllDays.xlsx"),
    # add more years here if you like, e.g.:
    2018: Path(r"/Users/melis/Desktop/09_30_ndvi_excels/myadjustmenetmanually/NDVI_2018_AllDays.xlsx"),
    2019: Path(r"/Users/melis/Desktop/09_30_ndvi_excels/myadjustmenetmanually/NDVI_2018_AllDays.xlsx"),
    2021: Path(r"/Users/melis/Desktop/09_30_ndvi_excels/myadjustmenetmanually/NDVI_2021_AllDays.xlsx"),
    2022: Path(r"/Users/melis/Desktop/09_30_ndvi_excels/myadjustmenetmanually/NDVI_2022_AllDays.xlsx"),
    2023: Path(r"/Users/melis/Desktop/09_30_ndvi_excels/myadjustmenetmanually/NDVI_2023_AllDays.xlsx"),
    2024: Path(r"/Users/melis/Desktop/09_30_ndvi_excels/myadjustmenetmanually/NDVI_2024_AllDays.xlsx"),
}
    
    
# Where to save outputs
OUT_DIR = Path("/Users/melis/Desktop/09_30_ndvi_excels")
OUT_DIR.mkdir(parents=True, exist_ok=True)

MASTER_JSON = OUT_DIR / "bbox_master_map.json"


# ------------------ HELPERS -----------------------

getcontext().prec = 28  # high precision for Decimal ops


def _to_dec_str(v, places=DECIMALS):
    """Normalize a number to a fixed-decimal string (e.g., '-124.372000')."""
    # Use Decimal on the ORIGINAL string to avoid float noise
    d = Decimal(str(v)).quantize(Decimal(f"1e-{places}"), rounding=ROUND_HALF_UP)
    # format exactly with 'places' decimals
    s = f"{d:.{places}f}"
    return s


def make_key_str(xmin, ymin, xmax, ymax, places=DECIMALS):
    """
    Make a canonical string key with fixed decimals:
      'xmin|ymin|xmax|ymax'
    """
    return "|".join([
        _to_dec_str(xmin, places),
        _to_dec_str(ymin, places),
        _to_dec_str(xmax, places),
        _to_dec_str(ymax, places),
    ])


def read_bbox_keys(excel_path):
    """
    Read only columns B:E (x_min, y_min, x_max, y_max).
    Tries by column labels first; falls back to 'B:E' if needed.
    Returns a list of canonical string keys and a DataFrame with those 4 cols.
    """
    try:
        df = pd.read_excel(excel_path, usecols=["x_min", "y_min", "x_max", "y_max"])
    except Exception:
        # Fallback to Excel letter range; assign names
        df = pd.read_excel(excel_path, usecols="B:E")
        df.columns = ["x_min", "y_min", "x_max", "y_max"]

    # Drop rows with any missing bbox values
    df = df.dropna(subset=["x_min", "y_min", "x_max", "y_max"]).copy()

    # Build normalized keys
    keys = [
        make_key_str(r.x_min, r.y_min, r.x_max, r.y_max)
        for r in df.itertuples(index=False)
    ]
    return keys, df

def save_year_csv(year, key_to_id, out_path):
    """
    Save per-year mapping as CSV with original numeric columns + ID.
    key_to_id: dict[str_key -> id]
    """
    rows = []
    for k, the_id in key_to_id.items():
        xmin, ymin, xmax, ymax = (float(s) for s in k.split("|"))
        rows.append({
            "x_min": xmin, "y_min": ymin, "x_max": xmax, "y_max": ymax, "ID": the_id
        })
    pd.DataFrame(rows).to_csv(out_path, index=False)



# --------------- CORE LOGIC -----------------------

def build_base_from_2021(path_2021: Path):
    keys_2021, df21 = read_bbox_keys(path_2021)

    total_rows = len(keys_2021)
    unique_keys = list(dict.fromkeys(keys_2021))  # keep order, drop dups
    n_unique = len(unique_keys)
    dups = total_rows - n_unique

    print(f"[2021] rows: {total_rows:,} | unique bboxes: {n_unique:,} | duplicates: {dups:,}")

    # IDs 1..N_2021
    master_map = {k: i+1 for i, k in enumerate(unique_keys)}
    next_id = n_unique + 1

    # Save 2021 CSV
    save_year_csv(2021, master_map, OUT_DIR / "bbox_ids_2021.csv")

    # Sanity: IDs are contiguous
    assigned = sorted(master_map.values())
    assert assigned == list(range(1, n_unique+1)), "Base ID sequence is not contiguous!"

    return master_map, next_id, n_unique, dups

def add_year(year: int, path: Path, master_map: dict, next_id_start: int):
    """Add one year's keys to the master map; reuse IDs or assign new ones."""
    keys, df = read_bbox_keys(path)
    total_rows = len(keys)
    unique_ordered = list(dict.fromkeys(keys))
    n_unique = len(unique_ordered)
    dups = total_rows - n_unique

    matched = 0
    new_added = 0
    next_id = next_id_start

    # Per-year view: key -> id actually used for this year
    year_key_to_id = {}

    for k in unique_ordered:
        if k in master_map:
            year_key_to_id[k] = master_map[k]
            matched += 1
        else:
            master_map[k] = next_id
            year_key_to_id[k] = next_id
            next_id += 1
            new_added += 1

    # Sanity checks
    assert matched + new_added == n_unique, "Accounting mismatch (matched + new != unique)."
    print(f"[{year}] rows: {total_rows:,} | unique: {n_unique:,} | dups: {dups:,} | "
          f"matched to master: {matched:,} | new IDs: {new_added:,} | next_id now: {next_id}")

    # Save per-year CSV
    save_year_csv(year, year_key_to_id, OUT_DIR / f"bbox_ids_{year}.csv")

    return next_id, matched, new_added, dups


# ---------------------- RUN -----------------------

if __name__ == "__main__":
    # 1) Seed from 2021
    master_map, next_id, n21_unique, n21_dups = build_base_from_2021(BASE_2021)

    # 2) Add other years (order matters for reporting, not for correctness)
    for yr in sorted(OTHER_YEARS.keys()):
        next_id, matched, new_added, dups = add_year(yr, OTHER_YEARS[yr], master_map, next_id)

    # 3) Save master map
    with open(MASTER_JSON, "w") as f:
        json.dump(master_map, f, indent=2)

    print(f"\n✅ Done. Master map size: {len(master_map):,} keys. "
          f"Saved to: {MASTER_JSON}")
    print(f"Per-year CSVs written under: {OUT_DIR}")


#%%



import json
import pandas as pd
from pathlib import Path
from decimal import Decimal, ROUND_HALF_UP

# === CONFIG ===
MASTER_JSON = Path("/Users/melis/Desktop/09_30_ndvi_excels/bbox_master_map.json")
EXCEL_DIR = Path("/Users/melis/Desktop/09_30_ndvi_excels/myadjustmenetmanually")

YEARS = [2017, 2018, 2019, 2020, 2021, 2022, 2023, 2024]

# === HELPER for bbox key formatting ===
def _to_dec_str(v, places=10):
    d = Decimal(str(v)).quantize(Decimal(f"1e-{places}"), rounding=ROUND_HALF_UP)
    return f"{d:.{places}f}"

def make_key(xmin, ymin, xmax, ymax):
    return "|".join([_to_dec_str(xmin), _to_dec_str(ymin), _to_dec_str(xmax), _to_dec_str(ymax)])

# === LOAD MASTER MAP ===
with open(MASTER_JSON, "r") as f:
    bbox_map = json.load(f)

# Invert mapping: ID → bbox_key
id_to_key = {v: k for k, v in bbox_map.items()}

# Initialize master NDVI dict
ndvi_master = {str(i): {"key": k, "ndvi": {}} for i, k in id_to_key.items()}

# === LOOP YEARS ===
for year in YEARS:
    excel_path = EXCEL_DIR / f"NDVI_{year}_AllDays.xlsx"
    if not excel_path.exists():
        print(f"⚠️ Skipping {year}, file not found")
        continue

    df = pd.read_excel(excel_path)
    df = df.dropna(subset=["x_min", "y_min", "x_max", "y_max"])

    # Build keys for this year
    df["key"] = df.apply(lambda r: make_key(r["x_min"], r["y_min"], r["x_max"], r["y_max"]), axis=1)

    # Extract day columns (all except first 5 metadata cols)
    day_cols = [c for c in df.columns if c not in ["ID", "x_min", "y_min", "x_max", "y_max", "key"]]

    for _, row in df.iterrows():
        k = row["key"]
        if k in bbox_map:
            idx = str(bbox_map[k])
            # Collect NDVI as dict {day_index: value}
            ndvi_values = {str(i): row[day] for i, day in enumerate(day_cols) if pd.notna(row[day])}
            ndvi_master[idx]["ndvi"][str(year)] = ndvi_values

# === SAVE MASTER JSON ===
out_path = MASTER_JSON.parent / "ndvi_master.json"
with open(out_path, "w") as f:
    json.dump(ndvi_master, f, indent=2)

print(f"✅ Saved NDVI master JSON: {out_path}")


#%%

#!/usr/bin/env python3
import json
import pandas as pd
from pathlib import Path

# === INPUT / OUTPUT ===
json_path = Path("/Users/melis/Desktop/09_30_ndvi_excels/ndvi_master.json")
out_path = Path("/Users/melis/Desktop/09_30_ndvi_excels/Merged_NDVI_AllYears.xlsx")

# === Load JSON ===
with open(json_path, "r", encoding="utf-8") as f:
    data = json.load(f)

rows = []
for id_str, entry in data.items():   # iterate id -> entry
    row = {
        "ID": int(id_str),
        "key": entry["key"],
    }

    # split bbox string into coordinates
    try:
        coords = entry["key"].split("|")   # we used '|' when saving keys
        row["x_min"] = float(coords[0])
        row["y_min"] = float(coords[1])
        row["x_max"] = float(coords[2])
        row["y_max"] = float(coords[3])
    except Exception:
        row["x_min"] = row["y_min"] = row["x_max"] = row["y_max"] = None

    # add NDVI values: each year contains {day: value}
    for year, ndvi_dict in entry["ndvi"].items():
        for day, val in ndvi_dict.items():
            row[int(day)] = val   # day index stays numeric

    rows.append(row)

df = pd.DataFrame(rows)

# === Order columns ===
day_cols = sorted([c for c in df.columns if isinstance(c, int)])
final_cols = ["ID", "x_min", "y_min", "x_max", "y_max"] + day_cols
df_final = df[final_cols]

out_path.parent.mkdir(parents=True, exist_ok=True)
df_final.to_excel(out_path, index=False, engine="openpyxl")

print(f"✅ Saved merged Excel: {out_path}")


#%% Just assign the correct id to coordintes in excel

#!/usr/bin/env python3
import json
import pandas as pd
from pathlib import Path
from decimal import Decimal, getcontext, ROUND_HALF_UP

# === CONFIG ===
DECIMALS = 10  # must match how bbox_master_map.json was generated

json_path = Path("/Users/melis/Desktop/09_30_ndvi_excels/bbox_master_map.json")
excel_dir = Path("/Users/melis/Desktop/09_30_ndvi_excels/myadjustmenetmanually")
out_dir   = excel_dir / "with_ids"
out_dir.mkdir(exist_ok=True)

getcontext().prec = 28

def _to_dec_str(v, places=DECIMALS):
    d = Decimal(str(v)).quantize(Decimal(f"1e-{places}"), rounding=ROUND_HALF_UP)
    return f"{d:.{places}f}"

def make_key_str(xmin, ymin, xmax, ymax, places=DECIMALS):
    return "|".join([
        _to_dec_str(xmin, places),
        _to_dec_str(ymin, places),
        _to_dec_str(xmax, places),
        _to_dec_str(ymax, places),
    ])

# === Load bbox map ===
with open(json_path, "r", encoding="utf-8") as f:
    bbox_map = json.load(f)  # {"xmin|ymin|xmax|ymax": ID}

# === Process each Excel ===
for f in sorted(excel_dir.glob("NDVI_*_AllDays.xlsx")):
    year = f.stem.split("_")[1]
    print(f"📂 Processing {year}...")

    df = pd.read_excel(f, dtype=object)
    
    # Ensure bbox columns exist (B:E in your sheets)
    df = df.rename(columns={
        df.columns[1]: "x_min",
        df.columns[2]: "y_min",
        df.columns[3]: "x_max",
        df.columns[4]: "y_max"
    })

    # Build canonical key and map to ID
    df["key"] = [
        make_key_str(xmin, ymin, xmax, ymax)
        for xmin, ymin, xmax, ymax in df[["x_min","y_min","x_max","y_max"]].itertuples(index=False, name=None)
    ]
    df["ID"] = df["key"].map(bbox_map)

    # Save new file with IDs
    out_file = out_dir / f"{f.stem}_withID.xlsx"
    df.to_excel(out_file, index=False, engine="openpyxl")
    print(f"   ✅ Saved: {out_file}")


#%% Merge the excels with IDs

#!/usr/bin/env python3
import pandas as pd
from pathlib import Path

# === CONFIG ===
excel_dir = Path("/Users/melis/Desktop/09_30_ndvi_excels/myadjustmenetmanually/with_ids")
out_path  = Path("/Users/melis/Desktop/09_30_ndvi_excels/Merged_NDVI_AllYears.xlsx")

# === Collect all yearly files ===
files = sorted(excel_dir.glob("NDVI_*_AllDays_withID.xlsx"))

master = None

for f in files:
    year = f.stem.split("_")[1]   # e.g. "2017"
    print(f"📂 Processing {year}...")

    df = pd.read_excel(f, dtype=object)

    # Keep ID + NDVI day columns (drop coords, key)
    cols = [c for c in df.columns if str(c).isdigit()]  # day columns (0,1,...)
    sub = df[["ID"] + cols].copy()

    # Rename day columns → year prefix
    sub = sub.rename(columns={c: f"{year}_{c}" for c in cols})

    # Ensure ID is int
    sub["ID"] = sub["ID"].astype(int)

    # Merge into master
    if master is None:
        master = sub
    else:
        sub = sub.set_index("ID")
        master = master.merge(sub, how="outer", left_index=True, right_index=True, validate="one_to_one")


# === Save final merged file ===
master = master.sort_index().reset_index()  # ID back as a column
master.to_excel(out_path, index=False, engine="openpyxl")


print(f"✅ Saved merged NDVI Excel with {master.shape[1]-1} NDVI columns: {out_path}")

#%%
f_2017= r"/Users/melis/Desktop/09_30_ndvi_excels/myadjustmenetmanually/with_ids/NDVI_2017_AllDays_withID.xlsx"

df_2017 = pd.read_excel(f_2017)

values_2017 = df_2017.iloc[:, 5:-2]

values_2017 = values_2017.astype(float)

ID_col_2017 = df_2017["ID"]

ndvi_dict_2017 = {
    int(ID_col_2017.iloc[i]): values_2017.iloc[i].tolist()
    for i in range(len(ID_col_2017))
}

# Example: check one entry
first_id = ID_col_2017.iloc[0]
print(f"ID {first_id} → {ndvi_dict_2017[first_id][:10]}")  # first 10 days
    

f_2018= r"/Users/melis/Desktop/09_30_ndvi_excels/myadjustmenetmanually/with_ids/NDVI_2018_AllDays_withID.xlsx"

df_2018 = pd.read_excel(f_2018)

values_2018 = df_2018.iloc[:, 5:-2]

values_2018 = values_2018.astype(float)

ID_col_2018 = df_2018["ID"]

ndvi_dict_2017 = {
    int(ID_col_2017.iloc[i]): values_2017.iloc[i].tolist()
    for i in range(len(ID_col_2017))
}    



from pathlib import Path

# Base path with your yearly excels
excel_dir = Path(r"/Users/melis/Desktop/09_30_ndvi_excels/myadjustmenetmanually/with_ids")

# Start with 2017 dictionary
ndvi_dict = ndvi_dict_2017

# Years to merge
years = [2018, 2019, 2020, 2021, 2022, 2023, 2024]

for yr in years:
    f = excel_dir / f"NDVI_{yr}_AllDays_withID.xlsx"
    print(f"📂 Processing {yr}...")

    df = pd.read_excel(f)

    # Extract NDVI values (all day columns)
    values = df.iloc[:, 5:-2].astype(float)
    ids = df["ID"]

    for i in range(len(ids)):
        row_id = int(ids.iloc[i])
        ndvi_vals = values.iloc[i].tolist()

        if row_id in ndvi_dict:
            # Extend existing time series
            ndvi_dict[row_id].extend(ndvi_vals)
        else:
            # Create new entry
            ndvi_dict[row_id] = ndvi_vals

print(f"✅ Merged NDVI dictionary with {len(ndvi_dict)} unique IDs")

#%% Corrected version


import pandas as pd
import numpy as np

# === Base year 2017 ===
f_2017 = r"/Users/melis/Desktop/09_30_ndvi_excels/myadjustmenetmanually/with_ids/NDVI_2017_AllDays_withID.xlsx"
df_2017 = pd.read_excel(f_2017)
values_2017 = df_2017.iloc[:, 5:-2].astype(float)
ID_col_2017 = df_2017["ID"]

ndvi_dict = {
    int(ID_col_2017.iloc[i]): values_2017.iloc[i].tolist()
    for i in range(len(ID_col_2017))
}

# Track global offset
offset = values_2017.shape[1]   # 365
print(f"2017 → days={offset}")

# === Add subsequent years ===
years = [2018, 2019, 2020, 2021, 2022, 2023, 2024]
for yr in years:
    f_year = fr"/Users/melis/Desktop/09_30_ndvi_excels/myadjustmenetmanually/with_ids/NDVI_{yr}_AllDays_withID.xlsx"
    df_year = pd.read_excel(f_year)
    values_year = df_year.iloc[:, 5:-2].astype(float)
    ID_col_year = df_year["ID"]
    n_days = values_year.shape[1]

    print(f"{yr} → days={n_days}, global offset={offset}")

    # Expand existing series with NaNs
    for k in ndvi_dict:
        ndvi_dict[k].extend([np.nan] * n_days)

    # Now overwrite with actual values for IDs present in this year
    for i in range(len(ID_col_year)):
        id_val = int(ID_col_year.iloc[i])
        if id_val not in ndvi_dict:
            ndvi_dict[id_val] = [np.nan] * offset + values_year.iloc[i].tolist()
        else:
            ndvi_dict[id_val][offset:offset+n_days] = values_year.iloc[i].tolist()

    offset += n_days  # move global pointer


import pickle

with open("ndvi_dict.pkl", "wb") as f:
    pickle.dump(ndvi_dict, f)

import json

with open("ndvi_dict.json", "w") as f:
    json.dump(ndvi_dict, f)
    
import pandas as pd

# Convert dict to DataFrame (IDs as columns, days as rows)
df = pd.DataFrame(ndvi_dict)

# Save to Excel
df.to_excel("ndvi_dict_transposed.xlsx", index=True, engine="openpyxl")

print("✅ Saved ndvi_dict_transposed.xlsx")


#%% Sanity check

import pandas as pd
import json
from pathlib import Path

# ---------------- Config ----------------
MASTER_JSON = Path("/Users/melis/Desktop/09_30_ndvi_excels/bbox_master_map.json")
YEARLY_CSVS = {
    2017: Path("/Users/melis/Desktop/09_30_ndvi_excels/bbox_ids_2017.csv"),
    2018: Path("/Users/melis/Desktop/09_30_ndvi_excels/bbox_ids_2018.csv"),
    2020: Path("/Users/melis/Desktop/09_30_ndvi_excels/bbox_ids_2020.csv"),
    2021: Path("/Users/melis/Desktop/09_30_ndvi_excels/bbox_ids_2021.csv"),
    2022: Path("/Users/melis/Desktop/09_30_ndvi_excels/bbox_ids_2022.csv"),
    2023: Path("/Users/melis/Desktop/09_30_ndvi_excels/bbox_ids_2023.csv"),
    2024: Path("/Users/melis/Desktop/09_30_ndvi_excels/bbox_ids_2024.csv"),
    # add more here...
}

# ---------------- Load Master ----------------
with open(MASTER_JSON, "r", encoding="utf-8") as f:
    master_map = json.load(f)

master_keys = set(master_map.keys())

# ---------------- Sanity Check ----------------
for year, csv_path in YEARLY_CSVS.items():
    df = pd.read_csv(csv_path)

    # Build keys from raw strings without rounding
    year_keys = set(
        df.apply(
            lambda r: f"{r['x_min']}|{r['y_min']}|{r['x_max']}|{r['y_max']}", axis=1
        )
    )

    missing = year_keys - master_keys
    if missing:
        print(f"[{year}] ⚠️ {len(missing)} keys in CSV not found in master map!")
        # Show a few samples
        for k in list(missing)[:5]:
            print("   Missing key:", k)
    else:
        print(f"[{year}] ✅ All {len(year_keys)} keys found in master map.")
